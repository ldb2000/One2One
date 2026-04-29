#!/usr/bin/env python3
"""
Parse la feuille `Backlog_2025` du fichier xlsx STTi_BACKLOG_PROJET_2026
et émet en stdout un JSON [{code,name,domain,phase,cp,at}, ...].

Usage :
    python3 import_projects_xlsx.py <path_to_xlsx>

Pré-requis : openpyxl (`pip3 install --user openpyxl`).

Mappings :
- Phase (col 4) : Build / Deploy → "Build" ; Cadrage / Design / Standby
  conservés tels quels ; Clos → ignoré.
- Domaine (col 1) → entité.
- CP (col 26) → chefDeProjet.
- AT (col 27) → architecte.
- Code projet (col 2) sert de clé d'unicité.

Le script ne touche pas à la base SwiftData ; c'est l'app qui fait les
upserts via ModelContext (évite tout conflit de verrou SQLite).
"""

import json
import sys
import openpyxl


def map_phase(value):
    if not value:
        return None
    p = str(value).strip()
    upper = p.upper()
    if upper in ("BUILD", "DEPLOY"):
        return "Build"
    if upper == "RUN":
        return "Run"
    if upper in ("CLOS", "CLOSE", "CLOSED", "CLÔTURÉ", "CLOTURE"):
        return None
    return p


def split_code_name(raw_name, code: str) -> str:
    name = (raw_name or code).strip()
    for sep in (" : ", ": ", " "):
        prefix = code + sep
        if name.startswith(prefix):
            return name[len(prefix):].strip()
    return name


def main(argv) -> int:
    if len(argv) < 2:
        print("usage: import_projects_xlsx.py <xlsx>", file=sys.stderr)
        return 2

    xlsx_path = argv[1]
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Backlog_2025" not in wb.sheetnames:
        print(json.dumps({"error": "Feuille 'Backlog_2025' absente"}), file=sys.stderr)
        return 1

    ws = wb["Backlog_2025"]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    for r in rows[14:]:
        if not r or not r[2]:
            continue
        code = str(r[2]).strip()
        name_full = r[3]
        domain = (str(r[1]).strip() if r[1] else "Sans domaine")
        phase = map_phase(r[4])
        if phase is None:
            continue
        cp = (str(r[26]).strip() if len(r) > 26 and r[26] else "")
        at_name = (str(r[27]).strip() if len(r) > 27 and r[27] else "")

        out.append({
            "code": code,
            "name": split_code_name(name_full, code),
            "domain": domain,
            "phase": phase,
            "cp": cp,
            "at": at_name,
        })

    json.dump(out, sys.stdout, ensure_ascii=False)
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
